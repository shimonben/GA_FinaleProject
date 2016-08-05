# Genetic Algorithm testing - created by:
#       Shimon Ben-Alul
#       Hila Fox

from openpyxl import load_workbook
import Steel
import Population
import GAutils

CONST_SEQUENCE_LENGTH = 20
CONST_POPULATION_SIZE = 100
CONST_GENERATIONS = 5000
CONST_MUTATION_PROBABILITY = 0.75
CONST_MAX_THICKNESS = 80
CONST_MAX_WIDTH = 10
CONST_MAX_ZINC_THICKNESS = 80
CONST_MAX_STEEL_GRADE = 10
CONST_MIN_THICKNESS = 20
CONST_MIN_WIDTH = 4
CONST_MIN_ZINC_THICKNESS = 40
CONST_MIN_STEEL_GRADE = 0.1


def main():
    wb = load_workbook('coils.xlsx')
    ws = wb.active
    coils = []
    i = 0
    for row in ws.iter_rows('B2:E21'):
        thickness = row[0].value
        width = row[1].value
        zinc_thickness = row[2].value
        steel_grade = row[3].value
        temp = Steel.Steel(thickness, width, zinc_thickness, steel_grade, i)
        i += 1
        coils.append(temp)
    testing_the_algorithm(coils)


def testing_the_algorithm(coils):
    population = Population.Population(coils)
    population.createInitial(CONST_POPULATION_SIZE)
    lst = []
    for i in range(CONST_GENERATIONS):
        population.updateGenesRange()
        selected = population.rouletteSelection()
        offspring = selected[0].crossover(selected[1])
        c1 = offspring[0]
        c2 = offspring[1]
        c1.mutate(CONST_MUTATION_PROBABILITY)
        c2.mutate(CONST_MUTATION_PROBABILITY)
        c1.evaluate(c1.getSequence(), Steel.calculate_max_penalty(), population.coils)
        c2.evaluate(c2.getSequence(), Steel.calculate_max_penalty(), population.coils)
        population.replacement_elitism(offspring[0], offspring[1])
        population.update_fitness()
        best = population.get_best_solution()
        print("Generation ", i, " best: ", population.get_chromosome_by_index(best[0]).getSequence(), ", with fitness: ", best[1])
        lst.append(best)
    return lst


if __name__ == "__main__": main()
