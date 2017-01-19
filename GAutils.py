from openpyxl import load_workbook
from openpyxl import Workbook
import Population
import Steel
from Operators.CrossOver import crossover
from Operators.Mutation import mutate
from Operators.Selection import rouletteSelection
from Operators.Replacement import replacement_elitism
import xlsxwriter
import main

CONST_SEQUENCE_LENGTH = main.CONST_COILS_IN_BATCH
CONST_POPULATION_SIZE = 1000
CONST_GENERATIONS = 1000
CONST_GENERATIONS_TO_TEST = 1000
CONST_MUTATION_PROBABILITY = 0.75
CONST_MAX_THICKNESS = 80
CONST_MAX_WIDTH = 10
CONST_MAX_ZINC_THICKNESS = 80
CONST_MAX_STEEL_GRADE = 10
CONST_MIN_THICKNESS = 20
CONST_MIN_WIDTH = 4
CONST_MIN_ZINC_THICKNESS = 40
CONST_MIN_STEEL_GRADE = 0.1
CONST_THRESHOLD = 0.5

CONST_THICK_DIF = CONST_MAX_THICKNESS - CONST_MIN_THICKNESS
CONST_ZINC_DIF = CONST_MAX_ZINC_THICKNESS - CONST_MIN_ZINC_THICKNESS
CONST_WIDTH_DIF = CONST_MAX_WIDTH - CONST_MIN_WIDTH
CONST_STEEL_DIF = CONST_MAX_STEEL_GRADE - CONST_MIN_STEEL_GRADE


def get_coils_from_excel():
    wb = load_workbook(main.CONST_EXCEL_FILE_NAME_TO_READ)
    ws = wb.active
    coils = []
    i = 0
    temp = 'B2:E'
    temp += str(CONST_SEQUENCE_LENGTH + 1)
    for row in ws.iter_rows(temp):
        if check_validity(row, i):
            thickness = row[3].value
            width = row[2].value
            zinc_thickness = row[1].value
            steel_grade = row[0].value
            temp = Steel.Steel(thickness, width, zinc_thickness, steel_grade, i)
            coils.append(temp)
            i += 1
        else:
            print("Values are not in range! program shut down")
            exit()
    return coils


def check_validity(row, i):
    if (row[3].value < CONST_MIN_THICKNESS) or (row[3].value > CONST_MAX_THICKNESS):
        print("in coil ", i + 1, ", there is a problem with thickness")
        return False
    if (row[2].value < CONST_MIN_WIDTH) or (row[2].value > CONST_MAX_WIDTH):
        print("in coil ", i + 1, ", there is a problem with width")
        return False
    if (row[1].value < CONST_MIN_ZINC_THICKNESS) or (row[1].value > CONST_MAX_ZINC_THICKNESS):
        print("in coil ", i + 1, ", there is a problem with zinc thickness")
        return False
    if (row[0].value < CONST_MIN_STEEL_GRADE) or (row[0].value > CONST_MAX_STEEL_GRADE):
        print("in coil ", i + 1, ", there is a problem with steel grade")
        return False
    return True


def testing_the_algorithm(coils):
    population = Population.Population(coils)
    population.createInitial(CONST_POPULATION_SIZE)
    lst = []
    temp = []
    sum_of_penalty = 0
    for i in range(CONST_GENERATIONS):
        import os
        os.system('cls')
        print("Pleas wait while for the algorithm to finish")
        print(str(int((i/CONST_GENERATIONS)*100))+"% finished")
        population.updateGenesRange()
        best = population.get_best_solution()
        if i == 0:
            temp.append(population.get_chromosome_by_index(best[0]).getSequence())
            temp.append(best[1])
            lst.append(temp)
            temp = []
        child_created = []
        for k in range(CONST_POPULATION_SIZE//2):
            selected = rouletteSelection(population)
            offspring = crossover(selected[0], selected[1])
            c1 = offspring[0]
            c2 = offspring[1]
            c1 = mutate(c1)
            c2 = mutate(c2)
            c1.evaluate(c1.getSequence(), Steel.calculate_max_penalty(), population.coils)
            c2.evaluate(c2.getSequence(), Steel.calculate_max_penalty(), population.coils)
            child_created.append(c1)
            child_created.append(c2)
        for k in range(CONST_POPULATION_SIZE//2):
            population = replacement_elitism(population, child_created[k*2], child_created[k*2+1])
        population.update_fitness()
        best = population.get_best_solution()
        if i == (CONST_GENERATIONS - 1):
            temp.append(population.get_chromosome_by_index(best[0]).getSequence())
            temp.append(best[1])
            lst.append(temp)
    best_seq = population.get_chromosome_by_index(best[0]).getSequence()
    transition = []
    for i in range(CONST_SEQUENCE_LENGTH - 1):
        temp = coils[best_seq[i]].calculate_penalty(coils[best_seq[i + 1]])
        sum_of_penalty += temp
        if temp > CONST_THRESHOLD:
            transition.append(1)
        else:
            transition.append(0)
    avg_of_penalty = sum_of_penalty/(CONST_SEQUENCE_LENGTH-1)
    save_data_to_excel(lst, transition, coils, avg_of_penalty)
    improvement_from_last_to_first = lst[1][1] - lst[0][1]
    return float((1 - float(lst[1][1])) / (1 - float(lst[0][1]))) * 100


def save_data_to_excel_first_and_last(lst):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "first generation:"
    ws["C1"] = "last generation:"
    ws["F1"] = "fitness improvement:"
    ws["F2"] = "penalty improvement:"
    ws["G1"] = float(float(lst[1][1]) / float(lst[0][1])) * 100
    ws["G2"] = float((1 - float(lst[1][1])) / (1 - float(lst[0][1]))) * 100
    cell = "A" + str(CONST_SEQUENCE_LENGTH + 1)
    ws[cell] = "first generation fit:"
    cell = "C" + str(CONST_SEQUENCE_LENGTH + 1)
    ws[cell] = "last generation fit:"
    cell = "B"
    for i in range(2):
        for j in range(CONST_SEQUENCE_LENGTH):
            temp = cell + str(j + 1)
            ws[temp] = int(lst[i][0][j])
        cell = "D"
    cell = "B" + str(CONST_SEQUENCE_LENGTH + 1)
    ws[cell] = float(lst[0][1])
    cell = "D" + str(CONST_SEQUENCE_LENGTH + 1)
    ws[cell] = float(lst[1][1])
    wb.save("1st & last gens.xlsx")


def testing_the_algorithm_1000_runs(coils):
    sum_first_generation_fit = 0
    sum_last_generation_fit = 0
    sum_fit_improvement = 0
    sum_penalty_improvement = 0
    temp_first = 0
    temp_last = 0
    for run in range(CONST_GENERATIONS_TO_TEST):
        population = Population.Population(coils)
        population.createInitial(CONST_POPULATION_SIZE)
        for i in range(CONST_GENERATIONS):
            population.updateGenesRange()
            best = population.get_best_solution()
            if i == 0:
                sum_first_generation_fit += best[1]
                temp_first = best[1]
            selected = rouletteSelection(population)
            offspring = crossover(selected[0], selected[1])
            c1 = offspring[0]
            c2 = offspring[1]
            c1 = mutate(c1)
            c2 = mutate(c2)
            c1.evaluate(c1.getSequence(), Steel.calculate_max_penalty(), population.coils)
            c2.evaluate(c2.getSequence(), Steel.calculate_max_penalty(), population.coils)
            population = replacement_elitism(population, offspring[0], offspring[1])
            population.update_fitness()
            best = population.get_best_solution()
            if i == (CONST_GENERATIONS - 1):
                sum_last_generation_fit += best[1]
                temp_last = best[1]
        sum_fit_improvement += (temp_last / temp_first) * 100
        sum_penalty_improvement += ((1 - temp_last) / (1 - temp_first)) * 100
        print(run)
    sum_first_generation_fit /= CONST_GENERATIONS_TO_TEST
    sum_last_generation_fit /= CONST_GENERATIONS_TO_TEST
    sum_fit_improvement /= CONST_GENERATIONS_TO_TEST
    sum_penalty_improvement /= CONST_GENERATIONS_TO_TEST
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "first generation avg:"
    ws["B1"] = "last generation avg:"
    ws["C1"] = "fitness improvement avg:"
    ws["D1"] = "penalty improvement avg:"
    ws["E1"] = "population size:"
    ws["A2"] = sum_first_generation_fit
    ws["B2"] = sum_last_generation_fit
    ws["C2"] = sum_fit_improvement
    ws["D2"] = sum_penalty_improvement
    ws["E2"] = CONST_POPULATION_SIZE
    file_name = str(CONST_GENERATIONS_TO_TEST) + " runs avg.xlsx"
    wb.save(file_name)


def save_data_to_excel(lst, transition, coils, avg_of_penalty):
    steel_grade_array = []
    zinc_thickness_array = []
    width_array = []
    thickness_array = []
    sequence_arr = []
    best_seq = lst[1][0]
    i = 1
    insertion_coils_for_penalty = 0
    counter = 1  # first headlines for the attributes
    for j in range(CONST_SEQUENCE_LENGTH):
        sequence_arr.append(best_seq[j] + 1)
        steel_grade_array.append(coils[best_seq[j]].steel_grade)
        zinc_thickness_array.append(coils[best_seq[j]].zinc_thickness)
        width_array.append(coils[best_seq[j]].width)
        thickness_array.append(coils[best_seq[j]].thickness)
        counter += 1  # row added
        if j < (CONST_SEQUENCE_LENGTH - 1) and transition[j] == 1:
            insertion_coils_for_penalty += 1
            steel_grade_array.append('')
            zinc_thickness_array.append('')
            width_array.append('')
            thickness_array.append('')
            sequence_arr.append("Transition")
            counter += 1  # row added for transition

    workbook = xlsxwriter.Workbook(main.CONST_EXCEL_FILE_NAME_TO_WRITE)
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({'bold': 1})
    headings = ['Sequence', 'Steel Grade', 'Zinc Thickness', 'Steel Width', 'Steel Thickness']
    worksheet.write_row('A1', headings, bold)
    worksheet.write_column('A2', sequence_arr)
    worksheet.write_column('B2', steel_grade_array)
    worksheet.write_column('C2', zinc_thickness_array)
    worksheet.write_column('D2', width_array)
    worksheet.write_column('E2', thickness_array)

    cell = "B" + str(CONST_SEQUENCE_LENGTH + 1 + insertion_coils_for_penalty + i)
    worksheet.write(cell, float(lst[1][1]))
    cell = "A" + str(CONST_SEQUENCE_LENGTH + 1 + insertion_coils_for_penalty + i)
    worksheet.write(cell, "last generation fit:")
    #cell = "A" + str(CONST_SEQUENCE_LENGTH + 2 + insertion_coils_for_penalty + i)
    #worksheet.write(cell, "penalty improvement:")
    #cell = "B" + str(CONST_SEQUENCE_LENGTH + 2 + insertion_coils_for_penalty + i)
    #worksheet.write(cell, float((1 - float(lst[1][1])) / (1 - float(lst[0][1]))) * 100)
    #cell = "B" + str(CONST_SEQUENCE_LENGTH + 3 + insertion_coils_for_penalty + i)
    #worksheet.write(cell, avg_of_penalty)
    #cell = "A" + str(CONST_SEQUENCE_LENGTH + 3 + insertion_coils_for_penalty + i)
    #worksheet.write(cell, "avg penalty:")

    chart3 = workbook.add_chart({'type': 'column'})
    chart3.add_series({
        'name': '=Sheet1!$B$1',
        'categories': '=Sheet1!$A$2:$A$' + str(counter),
        'values': '=Sheet1!$B$2:$B$' + str(counter),
    })
    chart3.set_title({'name': 'Steel Grade ('+str(Steel.CONST_STEEL_GRADE_PENALTY)+')'})
    chart3.set_x_axis({'name': 'Sequence'})
    chart3.set_y_axis({'name': 'Steel Grade'})
    chart3.set_style(3)
    worksheet.insert_chart('F12', chart3, {'x_scale': 2, 'y_scale': 0.75})

    chart2 = workbook.add_chart({'type': 'column'})
    chart2.add_series({
        'name': '=Sheet1!$C$1',
        'categories': '=Sheet1!$A$2:$A$' + str(counter),
        'values': '=Sheet1!$C$2:$C$' + str(counter),
    })
    chart2.set_title({'name': 'Zinc Thickness ('+str(Steel.CONST_ZINC_THICKNESS_PENALTY)+')'})
    chart2.set_x_axis({'name': 'Sequence'})
    chart2.set_y_axis({'name': 'Zinc Thickness'})
    chart2.set_style(3)
    worksheet.insert_chart('F1', chart2, {'x_scale': 2, 'y_scale': 0.75})

    chart1 = workbook.add_chart({'type': 'column'})
    chart1.add_series({
        'name': '=Sheet1!$D$1',
        'categories': '=Sheet1!$A$2:$A$' + str(counter),
        'values': '=Sheet1!$D$2:$D$' + str(counter),
    })
    chart1.set_title({'name': 'Steel Width (' + str(Steel.CONST_WIDTH_PENALTY) + ')'})
    chart1.set_x_axis({'name': 'Sequence'})
    chart1.set_y_axis({'name': 'Steel Width'})
    chart1.set_style(3)
    worksheet.insert_chart('F23', chart1, {'x_scale': 2, 'y_scale': 0.75})

    chart4 = workbook.add_chart({'type': 'column'})
    chart4.add_series({
        'name': '=Sheet1!$E$1',
        'categories': '=Sheet1!$A$2:$A$' + str(counter),
        'values': '=Sheet1!$E$2:$E$' + str(counter),
    })
    chart4.set_title({'name': 'Steel Thickness (' + str(Steel.CONST_THICKNESS_PENALTY) + ')'})
    chart4.set_x_axis({'name': 'Sequence'})
    chart4.set_y_axis({'name': 'Steel Thickness'})
    chart4.set_style(3)
    worksheet.insert_chart('F34', chart4, {'x_scale': 2, 'y_scale': 0.75})
    workbook.close()


def testing_the_algorithm_total_and_best_improvement(coils):
    population = Population.Population(coils)
    population.createInitial(CONST_POPULATION_SIZE)
    best_improve = []
    total_improve = []
    for i in range(CONST_GENERATIONS):
        population.updateGenesRange()
        selected = rouletteSelection(population)
        offspring = crossover(selected[0], selected[1])
        c1 = offspring[0]
        c2 = offspring[1]
        c1 = mutate(c1)
        c2 = mutate(c2)
        c1.evaluate(c1.getSequence(), Steel.calculate_max_penalty(), population.coils)
        c2.evaluate(c2.getSequence(), Steel.calculate_max_penalty(), population.coils)
        population = replacement_elitism(population, offspring[0], offspring[1])
        population.update_fitness()
        best = population.get_best_solution()
        best_improve.append(best[1])
        total_improve.append(population.getFitness())
    save_data_to_excel_self_and_total_improve(best_improve, total_improve)


def save_data_to_excel_self_and_total_improve(lst_best, lst_total):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "best solution improvement:"
    ws["B1"] = "population improvement:"
    cell1 = "A"
    cell2 = "B"
    for j in range(CONST_GENERATIONS):
        ws[cell1 + str(j + 2)] = float(lst_best[j])
        ws[cell2 + str(j + 2)] = float(lst_total[j])
    wb.save("total and best improvements.xlsx")
