import os

import openpyxl
from openpyxl import load_workbook
import Steel

CONST_SEQUENCE_LENGTH = 20

CONST_MAX_THICKNESS = 80
CONST_MAX_WIDTH = 10
CONST_MAX_ZINC_THICKNESS = 80
CONST_MAX_STEEL_GRADE = 10
CONST_MIN_THICKNESS = 20
CONST_MIN_WIDTH = 4
CONST_MIN_ZINC_THICKNESS = 40
CONST_MIN_STEEL_GRADE = 0.1


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "first generation:"
    ws["C1"] = "last generation:"
    cell = "B"
    for i in range(2):
        for j in range(CONST_SEQUENCE_LENGTH):
            temp = cell + str(j+1)
            ws[temp] = lst[i][j]
        cell = "D"
    wb.save("1st & last gens.xlsx")



if __name__ == "__main__":
    main()
