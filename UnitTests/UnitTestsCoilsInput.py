import os
import unittest
from openpyxl import load_workbook

CONST_MAX_THICKNESS = 80
CONST_MAX_WIDTH = 10
CONST_MAX_ZINC_THICKNESS = 80
CONST_MAX_STEEL_GRADE = 10
CONST_MIN_THICKNESS = 20
CONST_MIN_WIDTH = 4
CONST_MIN_ZINC_THICKNESS = 40
CONST_MIN_STEEL_GRADE = 0.1


class TestCoilsInput(unittest.TestCase):
    def test_if_excel_exist(self):
        self.assertFalse(os.path.isfile("co.xlsx"))
        self.assertTrue(os.path.isfile("coils fail.xlsx"))
        self.assertTrue(os.path.isfile("coils.xlsx"))

    def test_is_coils_legit(self):
        self.assertTrue(check_complete_excel_file("coils.xlsx"))
        self.assertFalse(check_complete_excel_file("coils fail.xlsx"))


def check_complete_excel_file(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
    for row in ws.iter_rows('B2:E21'):
        if (row[3].value < CONST_MIN_THICKNESS) or (row[3].value > CONST_MAX_THICKNESS):
            return False
        if (row[2].value < CONST_MIN_WIDTH) or (row[2].value > CONST_MAX_WIDTH):
            return False
        if (row[1].value < CONST_MIN_ZINC_THICKNESS) or (row[1].value > CONST_MAX_ZINC_THICKNESS):
            return False
        if (row[0].value < CONST_MIN_STEEL_GRADE) or (row[0].value > CONST_MAX_STEEL_GRADE):
            return False
    return True



if __name__ == '__main__':
    unittest.main()
